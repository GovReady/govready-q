from os import system
import pathlib
import pandas
from openpyxl import load_workbook
from rest_framework.decorators import action
from rest_framework.response import Response

from api.base.views.base import SerializerClasses
from api.base.views.viewsets import ReadOnlyViewSet
from api.controls.serializers.element import SimpleElementControlSerializer, DetailedElementControlSerializer
from api.controls.serializers.poam import DetailedPoamSerializer, SimplePoamSerializer, SimpleSpreadsheetPoamSerializer, SimpleUpdatePoamSpreadsheetSerializer
from api.controls.serializers.system import DetailedSystemSerializer, SimpleSystemSerializer, SystemCreateAndSetProposalSerializer, SystemRetrieveProposalsSerializer
from api.controls.serializers.system_assement_results import DetailedSystemAssessmentResultSerializer, \
    SimpleSystemAssessmentResultSerializer
from controls.models import System, Element, ElementControl, SystemAssessmentResult, Poam
from siteapp.models import Proposal, User


class SystemViewSet(ReadOnlyViewSet):
    queryset = System.objects.all()

    serializer_classes = SerializerClasses(retrieve=DetailedSystemSerializer,
                                           list=SimpleSystemSerializer,
                                           CreateAndSetProposal=SystemCreateAndSetProposalSerializer,
                                           retrieveProposals=SystemRetrieveProposalsSerializer,)

    @action(detail=True, url_path="CreateAndSetProposal", methods=["POST"])
    def CreateAndSetProposal(self, request, **kwargs):
        system, validated_data = self.validate_serializer_and_get_object(request)
        newProposal = Proposal.objects.create(
            user=User.objects.get(id=validated_data['userId']),
            requested_element=Element.objects.get(id=validated_data['elementId']),
            criteria_comment=validated_data['criteria_comment'],
            status=validated_data['status'],
        )
        newProposal.save()
        system.add_proposals([newProposal.id])
        system.save()
        serializer_class = self.get_serializer_class('retrieve')
        serializer = self.get_serializer(serializer_class, system)
        return Response(serializer.data)

    @action(detail=True, url_path="retrieveProposals", methods=["GET"])
    def retrieveProposals(self, request, **kwargs):
        system, validated_data = self.validate_serializer_and_get_object(request)
        system.save()

        serializer_class = self.get_serializer_class('retrieveProposals')
        serializer = self.get_serializer(serializer_class, system)
        return Response(serializer.data)
    
class SystemControlsViewSet(ReadOnlyViewSet):
    queryset = ElementControl.objects.all()

    serializer_classes = SerializerClasses(retrieve=DetailedElementControlSerializer,
                                           list=SimpleElementControlSerializer)

    NESTED_ROUTER_PKS = [{'pk': 'systems_pk', 'model_field': 'element.system'}]


class SystemAssessmentViewSet(ReadOnlyViewSet):
    queryset = SystemAssessmentResult.objects.all()

    serializer_classes = SerializerClasses(retrieve=DetailedSystemAssessmentResultSerializer,
                                           list=SimpleSystemAssessmentResultSerializer)

    NESTED_ROUTER_PKS = [{'pk': 'systems_pk', 'model_field': 'system'}]


class SystemPoamViewSet(ReadOnlyViewSet):
    queryset = Poam.objects.all()

    serializer_classes = SerializerClasses(retrieve=DetailedPoamSerializer,
                                           list=SimplePoamSerializer)

    NESTED_ROUTER_PKS = [{'pk': 'systems_pk', 'model_field': 'statement.consumer_element.system'}]

class SystemPoamSpreadsheetViewSet(ReadOnlyViewSet):
    queryset = System.objects.all()
    # queryset = System.objects.filter(pk=9)
    serializer_classes = SerializerClasses(
        retrieve=SimpleSpreadsheetPoamSerializer,
        list=SimpleSpreadsheetPoamSerializer,
        updateSpreadsheet=SimpleUpdatePoamSpreadsheetSerializer)

    @action(detail=True, url_path="updateSpreadsheet", methods=["PUT"])
    def updateSpreadsheet(self, request, **kwargs):
        system, validated_data = self.validate_serializer_and_get_object(request)
        for key, value in validated_data.items():
            if key == 'row':
                row = value
            if key == 'column':
                column = value
            if key == 'value':
                value = value
        print(row, column, value)

        system.save()
        fn = "local/poams_list.xlsx"
        #load excel file
        workbook = load_workbook(filename=fn)
        sheet = workbook.active
        
        #open workbook
        colNames = {}
        current = 0
        for col in sheet.iter_cols(min_row=2, max_row=2, min_col=1, max_col=sheet.max_column):
            colNames[col[0].value.lower()] = current
            current += 1
        
        #modify the desired cell
        sheet.cell(row=row+2, column=colNames[column]+1).value = value
        
        #save the file
        workbook.save(fn)

        serializer_class = self.get_serializer_class('updateSpreadsheet')
        serializer = self.get_serializer(serializer_class, system)
        return Response(serializer.data)
